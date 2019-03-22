import face_recognition
from PIL import Image, ImageDraw
import os
import pickle
from stat import *
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import time


# This is an example of running face recognition on a single image
# and drawing a box around each person that was identified.

# Load a sample picture and learn how to recognize it.

def detect_by_photo(filename):

    wb = load_workbook("reports.xlsx")
    sheet = wb['CSE15']

    max_row=sheet.max_row
    max_column=sheet.max_column

    known_face_encodings=[]
    known_face_names=[]

    currentDir = os.path.dirname(os.path.abspath(__file__))
    pklFolderPath=os.path.join(currentDir,'Pkl')

    for files in os.listdir(pklFolderPath):
        pathname=os.path.join(pklFolderPath,files)
        mode=os.stat(pathname)[ST_MODE]

        if S_ISDIR(mode):
            for pklfiles in os.listdir(pathname):
                with open(pklFolderPath+'\\'+files+'\\'+pklfiles,'rb') as pickle_load:
                    lst=pickle.load(pickle_load)
                    known_face_encodings.append(lst)
                    known_face_names.append(files)


                    
      
    currentDate=time.strftime("%d_%m_%y")

    colIndex=1
    for col_index in range(1,max_column+1):
        if sheet.cell(row=1,column=col_index).value==currentDate:
            colIndex=col_index
            break


    # Load an image with an unknown face
    unknown_image = face_recognition.load_image_file(filename)

    # Find all the faces and face encodings in the unknown image
    face_locations = face_recognition.face_locations(unknown_image)
    face_encodings = face_recognition.face_encodings(unknown_image, face_locations)

    # Convert the image to a PIL-format image so that we can draw on top of it with the Pillow library
    # See http://pillow.readthedocs.io/ for more about PIL/Pillow
    pil_image = Image.fromarray(unknown_image)
    # Create a Pillow ImageDraw Draw instance to draw with
    draw = ImageDraw.Draw(pil_image)

    # Loop through each face found in the unknown image
    for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
        # See if the face is a match for the known face(s)
        matches = face_recognition.compare_faces(known_face_encodings, face_encoding)

        name = "Unknown"

        # If a match was found in known_face_encodings, just use the first one.
        if True in matches:
            first_match_index = matches.index(True)
            name = known_face_names[first_match_index]
            
            for i in range(1,max_row+1):
                        #print('In Loop')
                        if str(sheet.cell(row=i,column=1).value) == name:
                            print('In for loop sheet    ',i,' Col index= ',colIndex)
                            #sheet.cell(row=2,column=1).value = 90
                            sheet.cell(row=i,column=colIndex).value = "Present"
                            break
                        else:
                            pass


        # Draw a box around the face using the Pillow module
        draw.rectangle(((left, top), (right, bottom)), outline=(0, 0, 255))

        # Draw a label with a name below the face
        text_width, text_height = draw.textsize(name)
        draw.rectangle(((left, bottom - text_height - 10), (right, bottom)), fill=(0, 0, 255), outline=(0, 0, 255))
        draw.text((left + 6, bottom - text_height - 5), name, fill=(255, 255, 255, 255))


    # Remove the drawing library from memory as per the Pillow docs
    del draw

    # Display the resulting image
    pil_image.show()

    # You can also save a copy of the new image to disk if you want by uncommenting this line
    pil_image.save("image_with_boxes.jpg")
    
    wb.save(filename='reports.xlsx')
