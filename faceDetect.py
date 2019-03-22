import os
import face_recognition
import cv2
import pickle
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import time
from stat import *


def face_detect():

    wb = load_workbook("reports.xlsx")
    sheet = wb['CSE15']

    max_row=sheet.max_row
    max_column=sheet.max_column

    known_face_encodings=[]
    known_face_names=[]

    currentDir = os.path.dirname(os.path.abspath(__file__))
    #imageFolderPath=os.path.join(currentDir,'Images')
    pklFolderPath=os.path.join(currentDir,'Pkl')

    for files in os.listdir(pklFolderPath):
        pathname=os.path.join(pklFolderPath,files)
        mode=os.stat(pathname)[ST_MODE]

        if S_ISDIR(mode):
            for pklfiles in os.listdir(pathname):

                with open(pklFolderPath+'\\'+files+'\\'+pklfiles,'rb') as pickle_load:
                    lst=pickle.load(pickle_load)
                    known_face_encodings.append(lst)
                    known_face_names.append(files)




    video_capture=cv2.VideoCapture(0)

    face_locations = []
    face_encodings = []
    face_names = []
    process_this_frame = True

    currentDate=time.strftime("%d_%m_%y")

    colIndex=1
    for col_index in range(1,max_column+1):
        if sheet.cell(row=1,column=col_index).value==currentDate:
            colIndex=col_index
            break

    #print(colIndex)
    while True:
        # Grab a single frame of video
        ret, frame = video_capture.read()

        # Resize frame of video to 1/4 size for faster face recognition processing
        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

        # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
        rgb_small_frame = small_frame[:, :, ::-1]

        # Only process every other frame of video to save time
        if process_this_frame:
            # Find all the faces and face encodings in the current frame of video
            face_locations = face_recognition.face_locations(rgb_small_frame)
            face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

            face_names = []
            for face_encoding in face_encodings:
                # See if the face is a match for the known face(s)
                matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
                name = "Unknown"

                # If a match was found in known_face_encodings, just use the first one.
                if True in matches:
                    first_match_index = matches.index(True)
                    name = known_face_names[first_match_index]
                    print(name)
                    for i in range(1,max_row+1):
                        #print('In Loop')
                        if str(sheet.cell(row=i,column=1).value) == name:
                            print('In for loop sheet    ',i,' Col index= ',colIndex)
                            #sheet.cell(row=2,column=1).value = 90
                            sheet.cell(row=i,column=colIndex).value = "Present"
                            break
                        else:
                            pass


                face_names.append(name)

        process_this_frame = not process_this_frame


        # Display the results
        for (top, right, bottom, left), name in zip(face_locations, face_names):
            # Scale back up face locations since the frame we detected in was scaled to 1/4 size
            top *= 4
            right *= 4
            bottom *= 4
            left *= 4

            # Draw a box around the face
            cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

            # Draw a label with a name below the face
            cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
            font = cv2.FONT_HERSHEY_DUPLEX
            cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)

        # Display the resulting image
        cv2.imshow('Video', frame)

        wb.save(filename='reports.xlsx')
        # Hit 'q' on the keyboard to quit!
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    # Release handle to the webcam
    video_capture.release()
    cv2.destroyAllWindows()


    



