from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import time
import os
import sqlite3

def making_spread_sheet():

    currentDate=time.strftime("%d_%m_%y")
    #currentDate='09_01_19'

    if(os.path.exists('./reports.xlsx')):
        wb=load_workbook(filename='reports.xlsx')
        sheet=wb['CSE15']

        for col_index in range(1,100):
            col=get_column_letter(col_index)

            if sheet.cell(row=1,column=col_index).value is None:
                sheet.cell(row=1,column=col_index).value=currentDate
                break
            elif sheet.cell(row=1,column=col_index).value == currentDate:
                break

        row_count=sheet.max_row
        sheetStudentsInfo=[]
        for i in range(3,row_count+1):
            cellRegd=sheet.cell(row=i,column=1)
            cellName=sheet.cell(row=i,column=2)
            tupleStudent=cellRegd.value,cellName.value
            sheetStudentsInfo.append(tupleStudent)

        conn=sqlite3.connect('test.db')
        curr=conn.cursor()
        curr.execute('SELECT * FROM CSED')
        databaseInfo=curr.fetchall()

        diff=[obj for obj in databaseInfo if obj not in sheetStudentsInfo]   
        
        print(diff)
        
        for info in diff:
            sheet.append(info)

        wb.save(filename='reports.xlsx')

    else:
        wb=Workbook()
        dest_filename='reports.xlsx'

        sheet=wb.active
        sheet.title="CSE15"
        sheet.append(('Regd No.','Name',currentDate))
        sheet.append(('','',''))

        #entering students name from pkl files

        conn=sqlite3.connect('test.db')
        curr=conn.cursor()
        curr.execute('SELECT * FROM CSED')
        allInfo=curr.fetchall()

        for row in allInfo:
            sheet.append(row)

        wb.save(dest_filename)